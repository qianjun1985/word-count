import customtkinter as ctk
from tkinter import filedialog, messagebox
import threading
import os
from pathlib import Path
from collections import Counter, defaultdict
from datetime import datetime
import re
import csv  # CSV 支持（Python 内置）

# PDF 支持
try:
    import pdfplumber
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False

# Excel 支持
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

# EPUB 支持
try:
    from ebooklib import epub
    EPUB_SUPPORT = True
except ImportError:
    EPUB_SUPPORT = False

ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

class WordProcessorGUI(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("英文单词处理工具 v1.5")
        self.geometry("1200x850")
        self.minsize(1000, 700)
        
        self.input_file = ctk.StringVar()
        self.output_dir = ctk.StringVar()
        self.output_format = ctk.StringVar(value="xlsx")
        self.sort_type = ctk.StringVar(value="frequency")
        self.show_freq = ctk.BooleanVar(value=True)
        self.grouped = ctk.BooleanVar(value=False)
        self.use_exclude = ctk.BooleanVar(value=True)
        
        self.exclude_words = set()
        
        self.create_widgets()
        self.load_default_exclude_words()
        
    def create_widgets(self):
        # 主容器 - 左右分栏
        main_paned = ctk.CTkFrame(self, corner_radius=0)
        main_paned.pack(fill="both", expand=True, padx=10, pady=10)
        
        # 左侧：主功能区域
        left_frame = ctk.CTkFrame(main_paned)
        left_frame.pack(side="left", fill="both", expand=True, padx=(0, 5))
        
        # 右侧：排除词表区域
        right_frame = ctk.CTkFrame(main_paned, width=350)
        right_frame.pack(side="right", fill="both", expand=False, padx=(5, 0))
        right_frame.pack_propagate(False)
        
        self.create_main_area(left_frame)
        self.create_exclude_area(right_frame)
    
    def create_main_area(self, parent):
        # 标题
        title_frame = ctk.CTkFrame(parent, fg_color="#2E86AB", corner_radius=10)
        title_frame.pack(fill="x", pady=(0, 10))
        
        title_label = ctk.CTkLabel(title_frame, text="📊 英文单词处理工具 v1.5", 
                                  font=ctk.CTkFont(size=28, weight="bold"), 
                                  text_color="white")
        title_label.pack(pady=20)
        
        # 文件选择
        file_frame = ctk.CTkFrame(parent)
        file_frame.pack(fill="x", pady=(0, 10))
        
        ctk.CTkLabel(file_frame, text="📁 文件选择", 
                    font=ctk.CTkFont(size=16, weight="bold")).pack(anchor="w", padx=10, pady=(10, 5))
        
        # 输入文件
        input_frame = ctk.CTkFrame(file_frame, fg_color="transparent")
        input_frame.pack(fill="x", padx=10, pady=5)
        ctk.CTkLabel(input_frame, text="输入文件:", width=80).pack(side="left")
        ctk.CTkEntry(input_frame, textvariable=self.input_file, width=500).pack(side="left", padx=10)
        ctk.CTkButton(input_frame, text="浏览...", command=self.browse_input, width=100).pack(side="left")
        
        # 输出目录
        output_frame = ctk.CTkFrame(file_frame, fg_color="transparent")
        output_frame.pack(fill="x", padx=10, pady=5)
        ctk.CTkLabel(output_frame, text="输出目录:", width=80).pack(side="left")
        ctk.CTkEntry(output_frame, textvariable=self.output_dir, width=500).pack(side="left", padx=10)
        ctk.CTkButton(output_frame, text="浏览...", command=self.browse_output, width=100).pack(side="left")
        
        # 选项配置
        option_frame = ctk.CTkFrame(parent)
        option_frame.pack(fill="x", pady=(0, 10))
        
        ctk.CTkLabel(option_frame, text="⚙️ 选项配置", 
                    font=ctk.CTkFont(size=16, weight="bold")).pack(anchor="w", padx=10, pady=(10, 5))
        
        # 输出格式（新增 CSV）
        format_frame = ctk.CTkFrame(option_frame, fg_color="transparent")
        format_frame.pack(fill="x", padx=10, pady=5)
        ctk.CTkLabel(format_frame, text="输出格式:", width=80).pack(side="left")
        ctk.CTkRadioButton(format_frame, text="Excel (.xlsx)", variable=self.output_format, 
                          value="xlsx", command=self.toggle_options).pack(side="left", padx=10)
        ctk.CTkRadioButton(format_frame, text="CSV (.csv)", variable=self.output_format, 
                          value="csv", command=self.toggle_options).pack(side="left", padx=10)
        ctk.CTkRadioButton(format_frame, text="文本 (.txt)", variable=self.output_format, 
                          value="txt", command=self.toggle_options).pack(side="left", padx=10)
        
        # 排序方式
        sort_frame = ctk.CTkFrame(option_frame, fg_color="transparent")
        sort_frame.pack(fill="x", padx=10, pady=5)
        ctk.CTkLabel(sort_frame, text="排序方式:", width=80).pack(side="left")
        ctk.CTkRadioButton(sort_frame, text="按词频排序", variable=self.sort_type, 
                          value="frequency").pack(side="left", padx=10)
        ctk.CTkRadioButton(sort_frame, text="按首字母排序", variable=self.sort_type, 
                          value="alphabet").pack(side="left", padx=10)
        
        # 其他选项
        option_list = ctk.CTkFrame(option_frame, fg_color="transparent")
        option_list.pack(fill="x", padx=10, pady=5)
        ctk.CTkCheckBox(option_list, text="显示词频统计", variable=self.show_freq).pack(side="left", padx=10)
        self.group_check = ctk.CTkCheckBox(option_list, text="按首字母分组输出", variable=self.grouped)
        self.group_check.pack(side="left", padx=10)
        
        # 处理按钮
        btn_frame = ctk.CTkFrame(parent, fg_color="transparent")
        btn_frame.pack(fill="x", pady=10)
        self.process_btn = ctk.CTkButton(btn_frame, text="🚀 开始处理", command=self.start_processing,
                                        font=ctk.CTkFont(size=18, weight="bold"),
                                        height=50, width=300)
        self.process_btn.pack()
        
        # 进度显示
        progress_frame = ctk.CTkFrame(parent)
        progress_frame.pack(fill="both", expand=True, pady=(0, 10))
        
        ctk.CTkLabel(progress_frame, text="📈 处理进度", 
                    font=ctk.CTkFont(size=16, weight="bold")).pack(anchor="w", padx=10, pady=(10, 5))
        
        self.progress_text = ctk.CTkTextbox(progress_frame, font=ctk.CTkFont(family="Consolas", size=11))
        self.progress_text.pack(fill="both", expand=True, padx=10, pady=5)
        
        self.progress_bar = ctk.CTkProgressBar(progress_frame)
        self.progress_bar.pack(fill="x", padx=10, pady=(0, 10))
        self.progress_bar.set(0)
        
        # 状态栏
        self.status_var = ctk.StringVar(value="  就绪")
        status_bar = ctk.CTkLabel(parent, textvariable=self.status_var, anchor="w", 
                                 fg_color="gray", height=30)
        status_bar.pack(side="bottom", fill="x")
        
        self.toggle_options()
        self.log("欢迎使用英文单词处理工具 v1.5！（作者：钱俊 | 西南大学外国语学院2004级英语本科/2011级英美文学硕士研究生）")
        self.log("✨ 新增支持：EPUB 格式输入、CSV 输出")
        self.log("✨ 新增功能：自定义排除词表（停用词）")
        if not PDF_SUPPORT:
            self.log("⚠️  pdfplumber 未安装")
        if not EXCEL_SUPPORT:
            self.log("⚠️  openpyxl 未安装")
        if not EPUB_SUPPORT:
            self.log("⚠️  ebooklib 未安装（EPUB 支持）")
    
    def create_exclude_area(self, parent):
        """创建排除词表区域"""
        title_label = ctk.CTkLabel(parent, text="🚫 排除词表", 
                                  font=ctk.CTkFont(size=18, weight="bold"))
        title_label.pack(pady=(10, 5))
        
        info_label = ctk.CTkLabel(parent, text="这些单词不会出现在输出结果中", 
                                 font=ctk.CTkFont(size=11), text_color="gray")
        info_label.pack(pady=(0, 10))
        
        exclude_frame = ctk.CTkFrame(parent)
        exclude_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        ctk.CTkLabel(exclude_frame, text="编辑排除单词（每行一个）:", 
                    font=ctk.CTkFont(size=12)).pack(anchor="w", padx=5, pady=5)
        
        self.exclude_text = ctk.CTkTextbox(exclude_frame, font=ctk.CTkFont(family="Consolas", size=11))
        self.exclude_text.pack(fill="both", expand=True, padx=5, pady=5)
        
        btn_frame = ctk.CTkFrame(parent)
        btn_frame.pack(fill="x", padx=10, pady=10)
        
        ctk.CTkButton(btn_frame, text="📂 载入文件", command=self.load_exclude_file, 
                     width=100).pack(side="left", padx=5)
        ctk.CTkButton(btn_frame, text="💾 保存文件", command=self.save_exclude_file, 
                     width=100).pack(side="left", padx=5)
        ctk.CTkButton(btn_frame, text="🔄 重置", command=self.reset_exclude_words, 
                     width=100, fg_color="#FF9800").pack(side="left", padx=5)
        
        self.exclude_count_label = ctk.CTkLabel(parent, text="当前排除词数：0", 
                                               font=ctk.CTkFont(size=11), text_color="green")
        self.exclude_count_label.pack(pady=5)
        
        self.exclude_check = ctk.CTkCheckBox(parent, text="启用排除功能", 
                                            variable=self.use_exclude)
        self.exclude_check.pack(pady=5)
        self.exclude_check.select()
    
    def load_default_exclude_words(self):
        default_excludes = [
            'the', 'a', 'an', 'i', 'you', 'he', 'she', 'it', 'we', 'they',
            'me', 'him', 'her', 'us', 'them', 'my', 'your', 'his', 'its', 'our', 'their',
            'this', 'that', 'these', 'those', 'in', 'on', 'at', 'to', 'for', 'of', 'with', 'by',
            'from', 'up', 'down', 'out', 'off', 'over', 'under', 'and', 'or', 'but', 'if',
            'because', 'as', 'while', 'is', 'am', 'are', 'was', 'were', 'be', 'been', 'being',
            'have', 'has', 'had', 'do', 'does', 'did', 'not', 'no', 'yes', 'so', 'than',
            'too', 'very', 'can', 'will', 'just', 'should', 'now',
        ]
        self.exclude_text.insert("0.0", "\n".join(default_excludes))
        self.update_exclude_count()
    
    def update_exclude_count(self):
        words = self.get_exclude_words()
        self.exclude_count_label.configure(text=f"当前排除词数：{len(words)}")
        self.exclude_words = set(words)
    
    def get_exclude_words(self):
        text = self.exclude_text.get("0.0", "end").strip()
        words = [w.strip().lower() for w in text.split('\n') if w.strip()]
        return words
    
    def load_exclude_file(self):
        filetypes = [("文本文件", "*.txt"), ("所有文件", "*.*")]
        filename = filedialog.askopenfilename(title="选择排除词表文件", filetypes=filetypes)
        if filename:
            try:
                with open(filename, 'r', encoding='utf-8') as f:
                    content = f.read()
                current = self.exclude_text.get("0.0", "end").strip()
                if current:
                    self.exclude_text.insert("end", "\n" + content)
                else:
                    self.exclude_text.insert("0.0", content)
                self.update_exclude_count()
                self.log(f"✓ 已载入排除词表：{filename}")
            except Exception as e:
                messagebox.showerror("错误", f"载入失败：\n{str(e)}")
    
    def save_exclude_file(self):
        filename = filedialog.asksaveasfilename(title="保存排除词表", 
                                               defaultextension=".txt",
                                               filetypes=[("文本文件", "*.txt")])
        if filename:
            try:
                content = self.exclude_text.get("0.0", "end").strip()
                with open(filename, 'w', encoding='utf-8') as f:
                    f.write(content)
                self.log(f"✓ 已保存排除词表：{filename}")
            except Exception as e:
                messagebox.showerror("错误", f"保存失败：\n{str(e)}")
    
    def reset_exclude_words(self):
        if messagebox.askyesno("确认", "确定要重置为默认排除词表吗？"):
            self.exclude_text.delete("0.0", "end")
            self.load_default_exclude_words()
            self.log("✓ 已重置为默认排除词表")
    
    def toggle_options(self):
        if self.output_format.get() == "xlsx":
            self.grouped.set(True)
            self.group_check.configure(state="disabled")
        else:
            self.group_check.configure(state="normal")
    
    def browse_input(self):
        # 新增 EPUB 格式
        filetypes = [("所有支持的文件", "*.pdf *.txt *.text *.md *.epub"), 
                    ("PDF 文件", "*.pdf"), 
                    ("EPUB 文件", "*.epub"),
                    ("文本文件", "*.txt *.text *.md"), 
                    ("所有文件", "*.*")]
        filename = filedialog.askopenfilename(title="选择输入文件", filetypes=filetypes)
        if filename:
            self.input_file.set(filename)
            self.log(f"✓ 已选择输入文件：{filename}")
            self.output_dir.set(os.path.dirname(filename))
    
    def browse_output(self):
        directory = filedialog.askdirectory(title="选择输出目录")
        if directory:
            self.output_dir.set(directory)
            self.log(f"✓ 已选择输出目录：{directory}")
    
    def log(self, message):
        self.progress_text.insert("end", message + "\n")
        self.progress_text.see("end")
        self.update_idletasks()
    
    def update_status(self, message):
        self.status_var.set(message)
    
    def update_progress(self, value):
        self.progress_bar.set(value)
    
    def start_processing(self):
        if not self.input_file.get():
            messagebox.showerror("错误", "请选择输入文件！")
            return
        if not self.output_dir.get():
            messagebox.showerror("错误", "请选择输出目录！")
            return
        if not os.path.exists(self.input_file.get()):
            messagebox.showerror("错误", "输入文件不存在！")
            return
        
        self.process_btn.configure(state="disabled", text="处理中...")
        self.progress_text.delete("1.0", "end")
        self.update_progress(0)
        
        thread = threading.Thread(target=self.process_files)
        thread.daemon = True
        thread.start()
    
    def process_files(self):
        try:
            input_file = self.input_file.get()
            output_dir = self.output_dir.get()
            output_format = self.output_format.get()
            use_exclude = self.use_exclude.get()
            
            self.log("=" * 60)
            self.log("开始处理文件...")
            self.update_progress(0.1)
            
            self.update_exclude_count()
            exclude_count = len(self.exclude_words)
            if use_exclude and exclude_count > 0:
                self.log(f"🚫 排除功能：启用 ({exclude_count} 个排除词)")
            else:
                self.log(f"🚫 排除功能：禁用")
            
            # 提取文本（新增 EPUB 支持）
            self.update_status("正在提取文本...")
            if input_file.lower().endswith('.pdf'):
                if not PDF_SUPPORT:
                    raise Exception("未安装 pdfplumber 库")
                text = self.extract_text_from_pdf(input_file)
            elif input_file.lower().endswith('.epub'):
                if not EPUB_SUPPORT:
                    raise Exception("未安装 ebooklib 库")
                text = self.extract_text_from_epub(input_file)
            else:
                text = self.extract_text_from_txt(input_file)
            
            self.update_progress(0.3)
            self.log(f"✓ 文本提取完成")
            
            # 提取单词
            self.update_status("正在提取单词...")
            words = self.extract_words(text)
            if not words:
                raise Exception("未找到任何英文单词")
            
            self.update_progress(0.5)
            original_count = len(words)
            self.log(f"✓ 提取到 {original_count} 个单词")
            
            # 应用排除词表
            if use_exclude and self.exclude_words:
                self.update_status("正在应用排除词表...")
                words_before = len(words)
                words = [w for w in words if w.lower() not in self.exclude_words]
                words_after = len(words)
                excluded_count = words_before - words_after
                self.log(f"✓ 排除了 {excluded_count} 个单词")
                self.log(f"  剩余 {words_after} 个单词")
            
            # 生成排序数据
            self.update_status("正在生成排序数据...")
            words_by_freq, word_freq = self.process_words_by_frequency(words)
            words_by_alpha, _ = self.process_words_by_alphabet(words)
            
            self.update_progress(0.7)
            self.log(f"✓ 生成 {len(words_by_freq)} 个唯一单词")
            
            # 准备统计信息
            stats_info = self.prepare_stats(words, words_by_freq, word_freq, input_file, exclude_count if use_exclude else 0)
            
            # 确定输出文件
            base_name = os.path.splitext(os.path.basename(input_file))[0]
            if output_format == "xlsx":
                output_file = os.path.join(output_dir, f"{base_name}_words.xlsx")
            elif output_format == "csv":
                output_file = os.path.join(output_dir, f"{base_name}_words.csv")
            else:
                output_file = os.path.join(output_dir, f"{base_name}_words.txt")
            
            # 写入输出（新增 CSV 支持）
            self.update_status("正在写入输出文件...")
            if output_format == "xlsx":
                if not EXCEL_SUPPORT:
                    raise Exception("未安装 openpyxl 库")
                success = self.write_output_excel(words_by_freq, words_by_alpha, word_freq,
                                                 output_file, self.show_freq.get(), stats_info)
                if not success:
                    raise Exception("Excel 文件写入失败")
            elif output_format == "csv":
                self.write_output_csv(words_by_freq, word_freq, output_file, stats_info)
            else:
                sort_type = self.sort_type.get()
                processed_words = words_by_freq if sort_type == "frequency" else words_by_alpha
                self.write_output(processed_words, output_file, word_freq, 
                                 self.grouped.get(), self.show_freq.get(),
                                 stats_info, sort_type)
            
            self.update_progress(1.0)
            self.log(f"\n✓ 处理完成！输出文件：{output_file}")
            self.update_status("处理完成！")
            
            messagebox.showinfo("成功", f"处理完成！\n\n输出文件：{output_file}")
            
        except Exception as e:
            self.log(f"\n❌ 错误：{str(e)}")
            self.update_status("处理失败")
            messagebox.showerror("错误", f"处理失败：\n{str(e)}")
        finally:
            self.process_btn.configure(state="normal", text="🚀 开始处理")
    
    def prepare_stats(self, words, words_by_freq, word_freq, input_file, excluded_count=0):
        stats_info = {
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'input_file': str(Path(input_file).absolute()),
            'total_words': len(words) + excluded_count,
            'unique_words': len(words_by_freq),
            'removed_duplicates': len(words) + excluded_count - len(words_by_freq),
            'excluded_words': excluded_count,
            'unique_count': len(set(word.lower() for word in words)),
        }
        if word_freq:
            stats_info['most_frequent_word'] = max(word_freq, key=word_freq.get)
            stats_info['most_frequent_count'] = word_freq[stats_info['most_frequent_word']]
            stats_info['least_frequent_word'] = min(word_freq, key=word_freq.get)
            stats_info['least_frequent_count'] = word_freq[stats_info['least_frequent_word']]
            stats_info['average_frequency'] = sum(word_freq.values()) / len(word_freq)
            stats_info['top_10'] = sorted(word_freq.items(), key=lambda x: -x[1])[:10]
        return stats_info
    
    def extract_text_from_epub(self, input_file):
        """从 EPUB 文件提取文本"""
        text = ""
        book = epub.read_epub(input_file)
        for item in book.get_items():
            if item.get_type() == 9:  # 9 = HTML 内容
                content = item.get_content().decode('utf-8', errors='ignore')
                # 清理 HTML 标签
                content = re.sub(r'<[^>]+>', ' ', content)
                text += content + "\n"
        return text
    
    def extract_text_from_pdf(self, input_file):
        text = ""
        with pdfplumber.open(input_file) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
        return text
    
    def extract_text_from_txt(self, input_file):
        encodings = ['utf-8', 'gbk', 'gb2312', 'latin-1', 'cp1252']
        for encoding in encodings:
            try:
                with open(input_file, 'r', encoding=encoding) as f:
                    return f.read()
            except UnicodeDecodeError:
                continue
        raise Exception("无法用常见编码读取文件")
    
    def extract_words(self, text):
        return re.findall(r"[a-zA-Z]+(?:[-'][a-zA-Z]+)*", text)
    
    def process_words_by_frequency(self, words):
        words_lower = [word.lower() for word in words]
        word_counter = Counter(words_lower)
        sorted_words = sorted(word_counter.items(), key=lambda x: (-x[1], x[0]))
        unique_words = [word for word, count in sorted_words]
        return unique_words, dict(sorted_words)
    
    def process_words_by_alphabet(self, words):
        words_lower = [word.lower() for word in words]
        unique_words = list(set(words_lower))
        unique_words.sort(key=lambda x: x[0].lower())
        return unique_words, None
    
    def group_by_first_letter(self, words, word_freq=None):
        groups = defaultdict(list)
        for word in words:
            if word:
                first_letter = word[0].lower()
                groups[first_letter].append(word)
        if word_freq:
            for letter in groups:
                groups[letter].sort(key=lambda x: -word_freq.get(x, 0))
        return dict(sorted(groups.items()))
    
    def write_output_csv(self, words, word_freq, output_file, stats_info=None):
        """写入 CSV 文件"""
        try:
            with open(output_file, 'w', encoding='utf-8-sig', newline='') as f:
                writer = csv.writer(f)
                
                # 写入统计信息（作为注释）
                writer.writerow(['# 单词处理统计报告'])
                writer.writerow(['#', f'生成时间：{stats_info.get("timestamp", "")}'])
                writer.writerow(['#', f'输入文件：{stats_info.get("input_file", "")}'])
                writer.writerow(['#', f'原始单词数：{stats_info.get("total_words", 0)}'])
                writer.writerow(['#', f'去重后单词数：{stats_info.get("unique_words", 0)}'])
                writer.writerow(['#', f'排除单词数：{stats_info.get("excluded_words", 0)}'])
                writer.writerow([])
                
                # 写入表头
                if self.show_freq.get():
                    writer.writerow(['序号', '单词', '出现次数'])
                else:
                    writer.writerow(['序号', '单词'])
                
                # 写入数据
                for i, word in enumerate(words, 1):
                    if self.show_freq.get():
                        writer.writerow([i, word, word_freq.get(word, 0)])
                    else:
                        writer.writerow([i, word])
            
            self.log(f"✓ 成功写入 CSV 文件")
            self.log(f"  单词总数：{len(words)}")
        except Exception as e:
            self.log(f"❌ CSV 写入错误：{e}")
            raise
    
    def write_output(self, words, output_file, word_freq=None, grouped=False, 
                    show_freq=False, stats_info=None, sort_type='frequency'):
        with open(output_file, 'w', encoding='utf-8') as f:
            if stats_info:
                f.write("=" * 60 + "\n")
                f.write("单词处理统计报告\n")
                f.write("=" * 60 + "\n\n")
                f.write(f"生成时间：{stats_info['timestamp']}\n")
                f.write(f"输入文件：{stats_info['input_file']}\n")
                f.write(f"输出文件：{output_file}\n")
                if stats_info.get('excluded_words', 0) > 0:
                    f.write(f"排除单词数：{stats_info['excluded_words']}\n")
                f.write("\n")
                f.write(f"原始单词数：{stats_info['total_words']}\n")
                f.write(f"去重后单词数：{stats_info['unique_words']}\n")
                f.write(f"删除重复数：{stats_info['removed_duplicates']}\n\n")
            
            if grouped and word_freq:
                groups = self.group_by_first_letter(words, word_freq)
                for letter in sorted(groups.keys()):
                    f.write(f"=== {letter.upper()} ===\n")
                    for word in groups[letter]:
                        if show_freq:
                            f.write(f"{word}\t({word_freq[word]})\n")
                        else:
                            f.write(f"{word}\n")
                    f.write("\n")
            else:
                for i, word in enumerate(words, 1):
                    if show_freq and word_freq:
                        f.write(f"{i}. {word}\t({word_freq[word]})\n")
                    else:
                        f.write(f"{i}. {word}\n")
    
    def write_output_excel(self, words_by_freq, words_by_alpha, word_freq, output_file,
                          show_freq=False, stats_info=None):
        try:
            self.log(f"🔍 开始写入 Excel...")
            self.log(f"  输出文件：{output_file}")
            self.log(f"  词频单词数：{len(words_by_freq)}")
            
            wb = Workbook()
            
            title_font = Font(bold=True, size=16, color="1F4E79")
            header_font = Font(bold=True, size=14, color="2E75B6")
            header_fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
            
            # 工作表 1：统计信息
            ws_stats = wb.active
            ws_stats.title = "📊 统计信息"
            
            ws_stats.cell(row=1, column=1, value="单词处理统计报告")
            ws_stats.cell(row=1, column=1).font = title_font
            ws_stats.cell(row=1, column=1).alignment = Alignment(horizontal='center')
            
            row = 3
            if stats_info:
                info_items = [
                    ("生成时间：", stats_info.get('timestamp', '')),
                    ("输入文件：", stats_info.get('input_file', '')),
                    ("输出文件：", output_file),
                ]
                if stats_info.get('excluded_words', 0) > 0:
                    info_items.append(("排除单词数：", stats_info.get('excluded_words', 0)))
                for label, value in info_items:
                    ws_stats.cell(row=row, column=1, value=label)
                    ws_stats.cell(row=row, column=2, value=value)
                    row += 1
                
                row += 1
                ws_stats.cell(row=row, column=1, value="基础统计")
                ws_stats.cell(row=row, column=1).font = header_font
                ws_stats.cell(row=row, column=1).fill = header_fill
                row += 1
                
                base_stats = [
                    ("原始单词数", stats_info.get('total_words', 0)),
                    ("去重后单词数", stats_info.get('unique_words', 0)),
                    ("删除重复数", stats_info.get('removed_duplicates', 0)),
                ]
                for label, value in base_stats:
                    ws_stats.cell(row=row, column=1, value=label)
                    ws_stats.cell(row=row, column=2, value=value)
                    row += 1
                
                if word_freq:
                    row += 1
                    ws_stats.cell(row=row, column=1, value="词频统计")
                    ws_stats.cell(row=row, column=1).font = header_font
                    ws_stats.cell(row=row, column=1).fill = header_fill
                    row += 1
                    
                    freq_stats = [
                        ("最高频单词", f"{stats_info.get('most_frequent_word', '')} ({stats_info.get('most_frequent_count', 0)} 次)"),
                        ("最低频单词", f"{stats_info.get('least_frequent_word', '')} ({stats_info.get('least_frequent_count', 0)} 次)"),
                        ("平均词频", f"{stats_info.get('average_frequency', 0):.2f} 次"),
                    ]
                    for label, value in freq_stats:
                        ws_stats.cell(row=row, column=1, value=label)
                        ws_stats.cell(row=row, column=2, value=value)
                        row += 1
                    
                    row += 1
                    ws_stats.cell(row=row, column=1, value="前 10 个高频单词")
                    ws_stats.cell(row=row, column=1).font = header_font
                    ws_stats.cell(row=row, column=1).fill = header_fill
                    row += 1
                    
                    for i, (word, count) in enumerate(stats_info.get('top_10', []), 1):
                        ws_stats.cell(row=row, column=1, value=f"{i}. {word}")
                        ws_stats.cell(row=row, column=2, value=f"{count} 次")
                        row += 1
            
            ws_stats.column_dimensions['A'].width = 20
            ws_stats.column_dimensions['B'].width = 45
            
            # 工作表 2：按词频排序
            ws_freq = wb.create_sheet(title="📈 按词频排序")
            
            freq_headers = ["序号", "单词", "出现次数"] if show_freq else ["序号", "单词"]
            ws_freq.append(freq_headers)
            
            for cell in ws_freq[1]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
            
            for i, word in enumerate(words_by_freq, 1):
                row_data = [i, word]
                if show_freq:
                    row_data.append(word_freq.get(word, 0))
                ws_freq.append(row_data)
            
            ws_freq.column_dimensions['A'].width = 10
            ws_freq.column_dimensions['B'].width = 30
            if show_freq:
                ws_freq.column_dimensions['C'].width = 15
            
            # 工作表 3：按首字母分组
            ws_group = wb.create_sheet(title="🔤 按首字母分组")
            
            group_headers = ["序号", "首字母", "单词", "出现次数"] if show_freq else ["序号", "首字母", "单词"]
            ws_group.append(group_headers)
            
            for cell in ws_group[1]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
            
            groups = self.group_by_first_letter(words_by_alpha, word_freq)
            
            global_index = 1
            for letter in sorted(groups.keys()):
                for word in groups[letter]:
                    row_data = [global_index, letter.upper(), word]
                    if show_freq:
                        row_data.append(word_freq.get(word, 0))
                    ws_group.append(row_data)
                    global_index += 1
            
            ws_group.column_dimensions['A'].width = 10
            ws_group.column_dimensions['B'].width = 12
            ws_group.column_dimensions['C'].width = 30
            if show_freq:
                ws_group.column_dimensions['D'].width = 15
            
            wb.save(output_file)
            self.log(f"✓ 成功写入 Excel 文件")
            self.log(f"  工作表：统计信息、按词频排序、按首字母分组")
            self.log(f"  单词总数：{len(words_by_freq)}")
            return True
            
        except Exception as e:
            self.log(f"❌ Excel 写入错误：{e}")
            import traceback
            traceback.print_exc()
            return False

def main():
    app = WordProcessorGUI()
    app.mainloop()

if __name__ == "__main__":
    main()