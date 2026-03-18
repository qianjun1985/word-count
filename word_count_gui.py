import customtkinter as ctk
from tkinter import filedialog, messagebox
import threading
import os
from pathlib import Path
from collections import Counter, defaultdict
from datetime import datetime
import re

# 尝试导入 PDF 和 Excel 库
try:
    import pdfplumber
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

# 设置 customtkinter 主题
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

class WordProcessorGUI(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("英文单词处理工具 v1.0")
        self.geometry("1000x800")
        self.minsize(800, 600)
        
        self.input_file = ctk.StringVar()
        self.output_dir = ctk.StringVar()
        self.output_format = ctk.StringVar(value="xlsx")
        self.sort_type = ctk.StringVar(value="frequency")
        self.show_freq = ctk.BooleanVar(value=True)
        self.grouped = ctk.BooleanVar(value=False)
        
        self.create_widgets()
        
    def create_widgets(self):
        # 主容器
        main_frame = ctk.CTkFrame(self, corner_radius=0)
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # 标题
        title_frame = ctk.CTkFrame(main_frame, fg_color="#2E86AB", corner_radius=10)
        title_frame.pack(fill="x", pady=(0, 10))
        title_label = ctk.CTkLabel(title_frame, text="📊 英文单词处理工具", 
                                  font=ctk.CTkFont(size=28, weight="bold"), 
                                  text_color="white")
        title_label.pack(pady=20)
        
        # 文件选择
        file_frame = ctk.CTkFrame(main_frame)
        file_frame.pack(fill="x", pady=(0, 10))
        ctk.CTkLabel(file_frame, text="📁 文件选择", 
                    font=ctk.CTkFont(size=16, weight="bold")).pack(anchor="w", padx=10, pady=(10, 5))
        
        input_frame = ctk.CTkFrame(file_frame, fg_color="transparent")
        input_frame.pack(fill="x", padx=10, pady=5)
        ctk.CTkLabel(input_frame, text="输入文件:", width=80).pack(side="left")
        ctk.CTkEntry(input_frame, textvariable=self.input_file, width=600).pack(side="left", padx=10)
        ctk.CTkButton(input_frame, text="浏览...", command=self.browse_input, width=100).pack(side="left")
        
        output_frame = ctk.CTkFrame(file_frame, fg_color="transparent")
        output_frame.pack(fill="x", padx=10, pady=5)
        ctk.CTkLabel(output_frame, text="输出目录:", width=80).pack(side="left")
        ctk.CTkEntry(output_frame, textvariable=self.output_dir, width=600).pack(side="left", padx=10)
        ctk.CTkButton(output_frame, text="浏览...", command=self.browse_output, width=100).pack(side="left")
        
        # 选项配置
        option_frame = ctk.CTkFrame(main_frame)
        option_frame.pack(fill="x", pady=(0, 10))
        ctk.CTkLabel(option_frame, text="⚙️ 选项配置", 
                    font=ctk.CTkFont(size=16, weight="bold")).pack(anchor="w", padx=10, pady=(10, 5))
        
        format_frame = ctk.CTkFrame(option_frame, fg_color="transparent")
        format_frame.pack(fill="x", padx=10, pady=5)
        ctk.CTkLabel(format_frame, text="输出格式:", width=80).pack(side="left")
        ctk.CTkRadioButton(format_frame, text="Excel (.xlsx)", variable=self.output_format, 
                          value="xlsx", command=self.toggle_options).pack(side="left", padx=10)
        ctk.CTkRadioButton(format_frame, text="文本 (.txt)", variable=self.output_format, 
                          value="txt", command=self.toggle_options).pack(side="left", padx=10)
        
        sort_frame = ctk.CTkFrame(option_frame, fg_color="transparent")
        sort_frame.pack(fill="x", padx=10, pady=5)
        ctk.CTkLabel(sort_frame, text="排序方式:", width=80).pack(side="left")
        ctk.CTkRadioButton(sort_frame, text="按词频排序", variable=self.sort_type, 
                          value="frequency").pack(side="left", padx=10)
        ctk.CTkRadioButton(sort_frame, text="按首字母排序", variable=self.sort_type, 
                          value="alphabet").pack(side="left", padx=10)
        
        option_list = ctk.CTkFrame(option_frame, fg_color="transparent")
        option_list.pack(fill="x", padx=10, pady=5)
        ctk.CTkCheckBox(option_list, text="显示词频统计", variable=self.show_freq).pack(side="left", padx=10)
        self.group_check = ctk.CTkCheckBox(option_list, text="按首字母分组输出", variable=self.grouped)
        self.group_check.pack(side="left", padx=10)
        
        # 处理按钮
        btn_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        btn_frame.pack(fill="x", pady=10)
        self.process_btn = ctk.CTkButton(btn_frame, text="🚀 开始处理", command=self.start_processing,
                                        font=ctk.CTkFont(size=18, weight="bold"),
                                        height=50, width=300)
        self.process_btn.pack()
        
        # 进度显示
        progress_frame = ctk.CTkFrame(main_frame)
        progress_frame.pack(fill="both", expand=True, pady=(0, 10))
        ctk.CTkLabel(progress_frame, text="📈 处理进度", 
                    font=ctk.CTkFont(size=16, weight="bold")).pack(anchor="w", padx=10, pady=(10, 5))
        self.progress_text = ctk.CTkTextbox(progress_frame, font=ctk.CTkFont(family="Consolas", size=11))
        self.progress_text.pack(fill="both", expand=True, padx=10, pady=5)
        self.progress_bar = ctk.CTkProgressBar(progress_frame)
        self.progress_bar.pack(fill="x", padx=10, pady=(0, 10))
        self.progress_bar.set(0)
        
        # 状态栏
        self.status_var = ctk.StringVar(value="就绪")
        status_bar = ctk.CTkLabel(self, textvariable=self.status_var, anchor="w", 
                                 fg_color="gray", height=30)
        status_bar.pack(side="bottom", fill="x")
        
        self.toggle_options()
        self.log("欢迎使用英文单词处理工具！（作者：钱俊）")
        if not PDF_SUPPORT:
            self.log("⚠️  警告：未安装 pdfplumber")
        if not EXCEL_SUPPORT:
            self.log("⚠️  警告：未安装 openpyxl")
    
    def toggle_options(self):
        if self.output_format.get() == "xlsx":
            self.grouped.set(True)
            self.group_check.configure(state="disabled")
        else:
            self.group_check.configure(state="normal")
    
    def browse_input(self):
        filetypes = [("所有支持的文件", "*.pdf *.txt *.text *.md"), ("PDF 文件", "*.pdf"), 
                    ("文本文件", "*.txt *.text *.md"), ("所有文件", "*.*")]
        filename = filedialog.askopenfilename(title="选择输入文件", filetypes=filetypes)
        if filename:
            self.input_file.set(filename)
            self.log(f"✓ 已选择输入文件：{filename}")
            self.output_dir.set(os.path.dirname(filename))
    
    def browse_output(self):
        directory = filedialog.askdirectory(title="选择输出目录")
        if directory:
            self.output_dir.set(directory)
            self.log(f"✓ 已选择输出目录：{directory}")
    
    def log(self, message):
        self.progress_text.insert("end", message + "\n")
        self.progress_text.see("end")
        self.update_idletasks()
    
    def update_status(self, message):
        self.status_var.set(message)
    
    def update_progress(self, value):
        self.progress_bar.set(value)
    
    def start_processing(self):
        if not self.input_file.get():
            messagebox.showerror("错误", "请选择输入文件！")
            return
        if not self.output_dir.get():
            messagebox.showerror("错误", "请选择输出目录！")
            return
        if not os.path.exists(self.input_file.get()):
            messagebox.showerror("错误", "输入文件不存在！")
            return
        
        self.process_btn.configure(state="disabled", text="处理中...")
        self.progress_text.delete("1.0", "end")
        self.update_progress(0)
        
        thread = threading.Thread(target=self.process_files)
        thread.daemon = True
        thread.start()
    
    def process_files(self):
        try:
            input_file = self.input_file.get()
            output_dir = self.output_dir.get()
            output_format = self.output_format.get()
            
            self.log("=" * 60)
            self.log("开始处理文件...")
            self.update_progress(0.1)
            
            # 提取文本
            self.update_status("正在提取文本...")
            if input_file.lower().endswith('.pdf'):
                if not PDF_SUPPORT:
                    raise Exception("未安装 pdfplumber 库")
                text = self.extract_text_from_pdf(input_file)
            else:
                text = self.extract_text_from_txt(input_file)
            
            self.update_progress(0.3)
            self.log(f"✓ 文本提取完成")
            
            # 提取单词
            self.update_status("正在提取单词...")
            words = self.extract_words(text)
            if not words:
                raise Exception("未找到任何英文单词")
            
            self.update_progress(0.5)
            self.log(f"✓ 提取到 {len(words)} 个单词")
            
            # 生成排序数据
            self.update_status("正在生成排序数据...")
            words_by_freq, word_freq = self.process_words_by_frequency(words)
            words_by_alpha, _ = self.process_words_by_alphabet(words)
            
            self.update_progress(0.7)
            self.log(f"✓ 生成 {len(words_by_freq)} 个唯一单词")
            
            # 准备统计信息
            stats_info = self.prepare_stats(words, words_by_freq, word_freq, input_file)
            
            # 确定输出文件
            base_name = os.path.splitext(os.path.basename(input_file))[0]
            if output_format == "xlsx":
                output_file = os.path.join(output_dir, f"{base_name}_words.xlsx")
            else:
                output_file = os.path.join(output_dir, f"{base_name}_words.txt")
            
            # 写入输出
            self.update_status("正在写入输出文件...")
            if output_format == "xlsx":
                if not EXCEL_SUPPORT:
                    raise Exception("未安装 openpyxl 库")
                success = self.write_output_excel(words_by_freq, words_by_alpha, word_freq,
                                                 output_file, self.show_freq.get(), stats_info)
                if not success:
                    raise Exception("Excel 文件写入失败")
            else:
                sort_type = self.sort_type.get()
                processed_words = words_by_freq if sort_type == "frequency" else words_by_alpha
                self.write_output(processed_words, output_file, word_freq, 
                                 self.grouped.get(), self.show_freq.get(),
                                 stats_info, sort_type)
            
            self.update_progress(1.0)
            self.log(f"\n✓ 处理完成！输出文件：{output_file}")
            self.update_status("处理完成！")
            
            messagebox.showinfo("成功", f"处理完成！\n\n输出文件：{output_file}")
            
        except Exception as e:
            self.log(f"\n❌ 错误：{str(e)}")
            self.update_status("处理失败")
            messagebox.showerror("错误", f"处理失败：\n{str(e)}")
        finally:
            self.process_btn.configure(state="normal", text="🚀 开始处理")
    
    def prepare_stats(self, words, words_by_freq, word_freq, input_file):
        stats_info = {
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'input_file': str(Path(input_file).absolute()),
            'total_words': len(words),
            'unique_words': len(words_by_freq),
            'removed_duplicates': len(words) - len(words_by_freq),
            'unique_count': len(set(word.lower() for word in words)),
        }
        if word_freq:
            stats_info['most_frequent_word'] = max(word_freq, key=word_freq.get)
            stats_info['most_frequent_count'] = word_freq[stats_info['most_frequent_word']]
            stats_info['least_frequent_word'] = min(word_freq, key=word_freq.get)
            stats_info['least_frequent_count'] = word_freq[stats_info['least_frequent_word']]
            stats_info['average_frequency'] = sum(word_freq.values()) / len(word_freq)
            stats_info['top_10'] = sorted(word_freq.items(), key=lambda x: -x[1])[:10]
        return stats_info
    
    def extract_text_from_pdf(self, input_file):
        text = ""
        with pdfplumber.open(input_file) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
        return text
    
    def extract_text_from_txt(self, input_file):
        encodings = ['utf-8', 'gbk', 'gb2312', 'latin-1', 'cp1252']
        for encoding in encodings:
            try:
                with open(input_file, 'r', encoding=encoding) as f:
                    return f.read()
            except UnicodeDecodeError:
                continue
        raise Exception("无法用常见编码读取文件")
    
    def extract_words(self, text):
        return re.findall(r"[a-zA-Z]+(?:[-'][a-zA-Z]+)*", text)
    
    def process_words_by_frequency(self, words):
        words_lower = [word.lower() for word in words]
        word_counter = Counter(words_lower)
        sorted_words = sorted(word_counter.items(), key=lambda x: (-x[1], x[0]))
        unique_words = [word for word, count in sorted_words]
        return unique_words, dict(sorted_words)
    
    def process_words_by_alphabet(self, words):
        words_lower = [word.lower() for word in words]
        unique_words = list(set(words_lower))
        unique_words.sort(key=lambda x: x[0].lower())
        return unique_words, None
    
    def group_by_first_letter(self, words, word_freq=None):
        groups = defaultdict(list)
        for word in words:
            if word:
                first_letter = word[0].lower()
                groups[first_letter].append(word)
        if word_freq:
            for letter in groups:
                groups[letter].sort(key=lambda x: -word_freq.get(x, 0))
        return dict(sorted(groups.items()))
    
    def write_output(self, words, output_file, word_freq=None, grouped=False, 
                    show_freq=False, stats_info=None, sort_type='frequency'):
        with open(output_file, 'w', encoding='utf-8') as f:
            if stats_info:
                f.write("=" * 60 + "\n")
                f.write("单词处理统计报告\n")
                f.write("=" * 60 + "\n\n")
                f.write(f"生成时间：{stats_info['timestamp']}\n")
                f.write(f"输入文件：{stats_info['input_file']}\n")
                f.write(f"输出文件：{output_file}\n\n")
                f.write(f"原始单词数：{stats_info['total_words']}\n")
                f.write(f"去重后单词数：{stats_info['unique_words']}\n")
                f.write(f"删除重复数：{stats_info['removed_duplicates']}\n\n")
            
            if grouped and word_freq:
                groups = self.group_by_first_letter(words, word_freq)
                for letter in sorted(groups.keys()):
                    f.write(f"=== {letter.upper()} ===\n")
                    for word in groups[letter]:
                        if show_freq:
                            f.write(f"{word}\t({word_freq[word]})\n")
                        else:
                            f.write(f"{word}\n")
                    f.write("\n")
            else:
                for i, word in enumerate(words, 1):
                    if show_freq and word_freq:
                        f.write(f"{i}. {word}\t({word_freq[word]})\n")
                    else:
                        f.write(f"{i}. {word}\n")
    
    def write_output_excel(self, words_by_freq, words_by_alpha, word_freq, output_file,
                          show_freq=False, stats_info=None):
        """完整的 Excel 写入函数（修复版）"""
        try:
            self.log(f"🔍 开始写入 Excel...")
            self.log(f"  输出文件：{output_file}")
            self.log(f"  词频单词数：{len(words_by_freq)}")
            self.log(f"  词频字典大小：{len(word_freq) if word_freq else 0}")
            
            wb = Workbook()
            
            title_font = Font(bold=True, size=16, color="1F4E79")
            header_font = Font(bold=True, size=14, color="2E75B6")
            header_fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
            
            # 工作表 1：统计信息
            ws_stats = wb.active
            ws_stats.title = "📊 统计信息"
            
            ws_stats.cell(row=1, column=1, value="单词处理统计报告")
            ws_stats.cell(row=1, column=1).font = title_font
            ws_stats.cell(row=1, column=1).alignment = Alignment(horizontal='center')
            
            row = 3
            if stats_info:
                info_items = [
                    ("生成时间：", stats_info.get('timestamp', '')),
                    ("输入文件：", stats_info.get('input_file', '')),
                    ("输出文件：", output_file),
                ]
                for label, value in info_items:
                    ws_stats.cell(row=row, column=1, value=label)
                    ws_stats.cell(row=row, column=2, value=value)
                    row += 1
                
                row += 1
                ws_stats.cell(row=row, column=1, value="基础统计")
                ws_stats.cell(row=row, column=1).font = header_font
                ws_stats.cell(row=row, column=1).fill = header_fill
                row += 1
                
                base_stats = [
                    ("原始单词数", stats_info.get('total_words', 0)),
                    ("去重后单词数", stats_info.get('unique_words', 0)),
                    ("删除重复数", stats_info.get('removed_duplicates', 0)),
                    ("唯一单词数", stats_info.get('unique_count', 0)),
                ]
                for label, value in base_stats:
                    ws_stats.cell(row=row, column=1, value=label)
                    ws_stats.cell(row=row, column=2, value=value)
                    row += 1
                
                if word_freq:
                    row += 1
                    ws_stats.cell(row=row, column=1, value="词频统计")
                    ws_stats.cell(row=row, column=1).font = header_font
                    ws_stats.cell(row=row, column=1).fill = header_fill
                    row += 1
                    
                    freq_stats = [
                        ("最高频单词", f"{stats_info.get('most_frequent_word', '')} ({stats_info.get('most_frequent_count', 0)} 次)"),
                        ("最低频单词", f"{stats_info.get('least_frequent_word', '')} ({stats_info.get('least_frequent_count', 0)} 次)"),
                        ("平均词频", f"{stats_info.get('average_frequency', 0):.2f} 次"),
                    ]
                    for label, value in freq_stats:
                        ws_stats.cell(row=row, column=1, value=label)
                        ws_stats.cell(row=row, column=2, value=value)
                        row += 1
                    
                    row += 1
                    ws_stats.cell(row=row, column=1, value="前 10 个高频单词")
                    ws_stats.cell(row=row, column=1).font = header_font
                    ws_stats.cell(row=row, column=1).fill = header_fill
                    row += 1
                    
                    for i, (word, count) in enumerate(stats_info.get('top_10', []), 1):
                        ws_stats.cell(row=row, column=1, value=f"{i}. {word}")
                        ws_stats.cell(row=row, column=2, value=f"{count} 次")
                        row += 1
            
            ws_stats.column_dimensions['A'].width = 20
            ws_stats.column_dimensions['B'].width = 45
            
            # 工作表 2：按词频排序
            ws_freq = wb.create_sheet(title="📈 按词频排序")
            
            freq_headers = ["序号", "单词", "出现次数"] if show_freq else ["序号", "单词"]
            ws_freq.append(freq_headers)
            
            for cell in ws_freq[1]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
            
            for i, word in enumerate(words_by_freq, 1):
                row_data = [i, word]
                if show_freq:
                    row_data.append(word_freq.get(word, 0))
                ws_freq.append(row_data)
            
            ws_freq.column_dimensions['A'].width = 10
            ws_freq.column_dimensions['B'].width = 30
            if show_freq:
                ws_freq.column_dimensions['C'].width = 15
            
            # 工作表 3：按首字母分组
            ws_group = wb.create_sheet(title="🔤 按首字母分组")
            
            group_headers = ["序号", "首字母", "单词", "出现次数"] if show_freq else ["序号", "首字母", "单词"]
            ws_group.append(group_headers)
            
            for cell in ws_group[1]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
            
            groups = self.group_by_first_letter(words_by_alpha, word_freq)
            
            global_index = 1
            for letter in sorted(groups.keys()):
                for word in groups[letter]:
                    row_data = [global_index, letter.upper(), word]
                    if show_freq:
                        row_data.append(word_freq.get(word, 0))
                    ws_group.append(row_data)
                    global_index += 1
            
            ws_group.column_dimensions['A'].width = 10
            ws_group.column_dimensions['B'].width = 12
            ws_group.column_dimensions['C'].width = 30
            if show_freq:
                ws_group.column_dimensions['D'].width = 15
            
            wb.save(output_file)
            self.log(f"✓ 成功写入 Excel 文件")
            self.log(f"  工作表：统计信息、按词频排序、按首字母分组")
            self.log(f"  单词总数：{len(words_by_freq)}")
            return True
            
        except Exception as e:
            self.log(f"❌ Excel 写入错误：{e}")
            import traceback
            traceback.print_exc()
            return False

def main():
    app = WordProcessorGUI()
    app.mainloop()

if __name__ == "__main__":
    main()